import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
import subprocess
import sys
import os
import threading
import webbrowser
from datetime import datetime

class GoogleDriveApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Google Drive File Manager")
        self.root.geometry("900x700")
        self.root.resizable(True, True)
        
        # Variables
        self.service = None
        self.current_folder_id = 'root'
        self.folder_stack = []
        self.files_data = []
        
        # Setup UI
        self.setup_ui()
        
    def setup_ui(self):
        # Main container
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(3, weight=1)
        
        # Title
        title_label = ttk.Label(main_frame, text="Google Drive File Manager", 
                                font=('Arial', 16, 'bold'))
        title_label.grid(row=0, column=0, pady=(0, 10), sticky=tk.W)
        
        # Button frame
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        
        self.install_btn = ttk.Button(button_frame, text="Install Dependencies", 
                                      command=self.install_dependencies)
        self.install_btn.grid(row=0, column=0, padx=(0, 5))
        
        self.login_btn = ttk.Button(button_frame, text="Google Sign In", 
                                    command=self.google_signin, state='disabled')
        self.login_btn.grid(row=0, column=1, padx=5)
        
        self.back_btn = ttk.Button(button_frame, text="← Back", 
                                   command=self.go_back, state='disabled')
        self.back_btn.grid(row=0, column=2, padx=5)
        
        self.refresh_btn = ttk.Button(button_frame, text="🔄 Refresh", 
                                      command=self.refresh_view, state='disabled')
        self.refresh_btn.grid(row=0, column=3, padx=5)
        
        self.export_btn = ttk.Button(button_frame, text="Export to Excel", 
                                     command=self.export_to_excel, state='disabled')
        self.export_btn.grid(row=0, column=4, padx=5)
        
        # Current path label
        self.path_label = ttk.Label(main_frame, text="Path: Not connected", 
                                    font=('Arial', 9))
        self.path_label.grid(row=2, column=0, sticky=tk.W, pady=(0, 5))
        
        # File list frame with scrollbar
        list_frame = ttk.Frame(main_frame)
        list_frame.grid(row=3, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        list_frame.columnconfigure(0, weight=1)
        list_frame.rowconfigure(0, weight=1)
        
        # Treeview for files
        columns = ('Name', 'Type', 'Size', 'Modified')
        self.tree = ttk.Treeview(list_frame, columns=columns, show='tree headings')
        
        self.tree.heading('#0', text='')
        self.tree.heading('Name', text='Name')
        self.tree.heading('Type', text='Type')
        self.tree.heading('Size', text='Size')
        self.tree.heading('Modified', text='Modified')
        
        self.tree.column('#0', width=30, stretch=False)
        self.tree.column('Name', width=350)
        self.tree.column('Type', width=100)
        self.tree.column('Size', width=100)
        self.tree.column('Modified', width=150)
        
        # Scrollbars
        vsb = ttk.Scrollbar(list_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(list_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        self.tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        vsb.grid(row=0, column=1, sticky=(tk.N, tk.S))
        hsb.grid(row=1, column=0, sticky=(tk.W, tk.E))
        
        # Bind double-click to folder navigation
        self.tree.bind('<Double-1>', self.on_item_double_click)
        
        # Status bar
        self.status_label = ttk.Label(main_frame, text="Ready. Click 'Install Dependencies' to begin.", 
                                      relief=tk.SUNKEN, anchor=tk.W)
        self.status_label.grid(row=4, column=0, sticky=(tk.W, tk.E), pady=(10, 0))
        
    def install_dependencies(self):
        """Install required Python packages"""
        self.status_label.config(text="Installing dependencies...")
        self.install_btn.config(state='disabled')
        
        def install():
            try:
                packages = [
                    'google-auth-oauthlib',
                    'google-auth-httplib2',
                    'google-api-python-client',
                    'openpyxl'
                ]
                
                for package in packages:
                    self.status_label.config(text=f"Installing {package}...")
                    subprocess.check_call([sys.executable, '-m', 'pip', 'install', package])
                
                self.root.after(0, lambda: self.status_label.config(text="Dependencies installed successfully!"))
                self.root.after(0, lambda: self.login_btn.config(state='normal'))
                self.root.after(0, lambda: messagebox.showinfo("Success", 
                    "All dependencies installed successfully!\nYou can now sign in to Google."))
            except Exception as e:
                self.root.after(0, lambda: self.status_label.config(text=f"Installation failed: {str(e)}"))
                self.root.after(0, lambda: messagebox.showerror("Error", 
                    f"Failed to install dependencies:\n{str(e)}"))
                self.root.after(0, lambda: self.install_btn.config(state='normal'))
        
        thread = threading.Thread(target=install, daemon=True)
        thread.start()
    
    def google_signin(self):
        """Authenticate with Google Drive"""
        self.status_label.config(text="Authenticating with Google...")
        self.login_btn.config(state='disabled')
        
        def authenticate():
            try:
                from google_auth_oauthlib.flow import InstalledAppFlow
                from google.auth.transport.requests import Request
                from googleapiclient.discovery import build
                import pickle
                
                SCOPES = ['https://www.googleapis.com/auth/drive.readonly']
                creds = None
                
                # Check for existing token
                if os.path.exists('token.pickle'):
                    with open('token.pickle', 'rb') as token:
                        creds = pickle.load(token)
                
                # If no valid credentials, let user log in
                if not creds or not creds.valid:
                    if creds and creds.expired and creds.refresh_token:
                        creds.refresh(Request())
                    else:
                        # Create credentials.json if it doesn't exist
                        if not os.path.exists('credentials.json'):
                            self.root.after(0, lambda: messagebox.showwarning("Setup Required",
                                "Please follow these steps:\n\n"
                                "1. Go to Google Cloud Console\n"
                                "2. Create a project and enable Google Drive API\n"
                                "3. Create OAuth 2.0 credentials (Desktop app)\n"
                                "4. Download the JSON file as 'credentials.json'\n"
                                "5. Place it in the same folder as this program\n\n"
                                "Opening Google Cloud Console..."))
                            webbrowser.open('https://console.cloud.google.com/apis/credentials')
                            self.root.after(0, lambda: self.login_btn.config(state='normal'))
                            return
                        
                        flow = InstalledAppFlow.from_client_secrets_file(
                            'credentials.json', SCOPES)
                        creds = flow.run_local_server(port=0)
                    
                    # Save credentials
                    with open('token.pickle', 'wb') as token:
                        pickle.dump(creds, token)
                
                self.service = build('drive', 'v3', credentials=creds)
                
                self.root.after(0, self.on_signin_success)
                
            except Exception as e:
                self.root.after(0, lambda: self.status_label.config(text=f"Authentication failed: {str(e)}"))
                self.root.after(0, lambda: messagebox.showerror("Error", 
                    f"Failed to authenticate:\n{str(e)}"))
                self.root.after(0, lambda: self.login_btn.config(state='normal'))
        
        thread = threading.Thread(target=authenticate, daemon=True)
        thread.start()
    
    def on_signin_success(self):
        """Called after successful sign-in"""
        self.status_label.config(text="Signed in successfully!")
        self.refresh_btn.config(state='normal')
        self.export_btn.config(state='normal')
        messagebox.showinfo("Success", "Signed in to Google Drive successfully!")
        self.load_folder('root')
    
    def load_folder(self, folder_id):
        """Load contents of a folder"""
        if not self.service:
            return
        
        self.status_label.config(text="Loading folder contents...")
        
        def load():
            try:
                # Clear current items
                self.root.after(0, lambda: self.tree.delete(*self.tree.get_children()))
                self.files_data = []
                
                # Query for files in folder
                query = f"'{folder_id}' in parents and trashed=false"
                results = self.service.files().list(
                    q=query,
                    pageSize=1000,
                    fields="files(id, name, mimeType, size, modifiedTime)",
                    orderBy="folder,name"
                ).execute()
                
                items = results.get('files', [])
                
                for item in items:
                    is_folder = item['mimeType'] == 'application/vnd.google-apps.folder'
                    icon = '📁' if is_folder else '📄'
                    type_str = 'Folder' if is_folder else 'File'
                    size_str = self.format_size(int(item.get('size', 0))) if not is_folder else '-'
                    modified = item.get('modifiedTime', '')
                    if modified:
                        modified = datetime.fromisoformat(modified.replace('Z', '+00:00')).strftime('%Y-%m-%d %H:%M')
                    
                    self.root.after(0, lambda i=item, ic=icon, t=type_str, s=size_str, m=modified: 
                        self.tree.insert('', 'end', text=ic, 
                                       values=(i['name'], t, s, m),
                                       tags=('folder' if i['mimeType'] == 'application/vnd.google-apps.folder' else 'file',),
                                       iid=i['id']))
                    
                    # Store for export
                    self.files_data.append({
                        'name': item['name'],
                        'type': type_str,
                        'size': item.get('size', 0) if not is_folder else 0,
                        'size_formatted': size_str,
                        'modified': modified,
                        'id': item['id'],
                        'is_folder': is_folder
                    })
                
                self.root.after(0, lambda: self.status_label.config(text=f"Loaded {len(items)} items"))
                
            except Exception as e:
                self.root.after(0, lambda: self.status_label.config(text=f"Error loading folder: {str(e)}"))
                self.root.after(0, lambda: messagebox.showerror("Error", f"Failed to load folder:\n{str(e)}"))
        
        thread = threading.Thread(target=load, daemon=True)
        thread.start()
    
    def on_item_double_click(self, event):
        """Handle double-click on item"""
        item = self.tree.selection()
        if not item:
            return
        
        item_id = item[0]
        tags = self.tree.item(item_id, 'tags')
        
        if 'folder' in tags:
            # Navigate into folder
            self.folder_stack.append(self.current_folder_id)
            self.current_folder_id = item_id
            self.back_btn.config(state='normal')
            
            # Update path
            folder_name = self.tree.item(item_id, 'values')[0]
            current_path = self.path_label.cget('text')
            if current_path == "Path: Not connected":
                new_path = f"Path: {folder_name}"
            else:
                new_path = f"{current_path} / {folder_name}"
            self.path_label.config(text=new_path)
            
            self.load_folder(item_id)
    
    def go_back(self):
        """Navigate to parent folder"""
        if self.folder_stack:
            self.current_folder_id = self.folder_stack.pop()
            self.load_folder(self.current_folder_id)
            
            # Update path
            current_path = self.path_label.cget('text')
            parts = current_path.split(' / ')
            if len(parts) > 1:
                new_path = ' / '.join(parts[:-1])
                self.path_label.config(text=new_path)
            else:
                self.path_label.config(text="Path: My Drive")
            
            if not self.folder_stack:
                self.back_btn.config(state='disabled')
    
    def refresh_view(self):
        """Refresh current folder"""
        self.load_folder(self.current_folder_id)
    
    def export_to_excel(self):
        """Export file list to Excel"""
        if not self.files_data:
            messagebox.showwarning("No Data", "No files to export. Please browse to a folder first.")
            return
        
        self.status_label.config(text="Exporting to Excel...")
        
        def export():
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
                
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Google Drive Files"
                
                # Headers
                headers = ['File Name', 'Type', 'Size (Bytes)', 'Size (Formatted)', 'Modified Date']
                ws.append(headers)
                
                # Style headers
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")
                
                # Add data
                for file_data in self.files_data:
                    ws.append([
                        file_data['name'],
                        file_data['type'],
                        file_data['size'],
                        file_data['size_formatted'],
                        file_data['modified']
                    ])
                
                # Adjust column widths
                ws.column_dimensions['A'].width = 50
                ws.column_dimensions['B'].width = 15
                ws.column_dimensions['C'].width = 15
                ws.column_dimensions['D'].width = 15
                ws.column_dimensions['E'].width = 20
                
                # Save file
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f'google_drive_files_{timestamp}.xlsx'
                wb.save(filename)
                
                self.root.after(0, lambda: self.status_label.config(text=f"Exported to {filename}"))
                self.root.after(0, lambda: messagebox.showinfo("Success", 
                    f"File list exported successfully!\n\nSaved as: {filename}"))
                
            except Exception as e:
                self.root.after(0, lambda: self.status_label.config(text=f"Export failed: {str(e)}"))
                self.root.after(0, lambda: messagebox.showerror("Error", 
                    f"Failed to export:\n{str(e)}"))
        
        thread = threading.Thread(target=export, daemon=True)
        thread.start()
    
    @staticmethod
    def format_size(size_bytes):
        """Format bytes to human-readable size"""
        if size_bytes == 0:
            return "0 B"
        
        size_names = ["B", "KB", "MB", "GB", "TB"]
        i = 0
        size = float(size_bytes)
        
        while size >= 1024.0 and i < len(size_names) - 1:
            size /= 1024.0
            i += 1
        
        return f"{size:.2f} {size_names[i]}"


def main():
    root = tk.Tk()
    app = GoogleDriveApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()